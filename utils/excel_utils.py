import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from utils.time_utils import parse_created_time

BASE_DIR = Path(__file__).resolve().parents[2]
EXPORTS_DIR = BASE_DIR / "data" / "exports"
DEFAULT_FONT_NAME = "Calibri"
DEFAULT_FONT_SIZE = 11
EXCEL_CREATED_TIME_FORMAT = "%d/%m/%Y %H:%M:%S"
EXCEL_MAX_COLUMN_WIDTH = 255
LINK_COLUMN_MIN_WIDTH = 80


def safe_filename(text: str) -> str:
    """
    Remove all invalid characters for Windows filename
    """
    return re.sub(r'[\\/*?:"<>|]', "_", str(text))


def extract_group_name(group_input: str) -> str:
    """
    Extract meaningful group name or id from input
    """
    if not group_input:
        return "unknown"

    group_input = str(group_input)

    # lấy id hoặc slug sau /groups/
    match = re.search(r"/groups/([^/?]+)", group_input)
    if match:
        return match.group(1)

    return group_input


def sanitize_excel_message(message: str) -> str:
    normalized_message = str(message or "")
    normalized_message = normalized_message.replace("\\N", "\n").replace("\\n", "\n")
    normalized_message = normalized_message.replace("\r\n", "\n").replace("\r", "\n")
    normalized_message = re.sub(r"\n+", ". ", normalized_message)
    normalized_message = re.sub(r"\s{2,}", " ", normalized_message)
    return normalized_message.strip(" .")


def format_excel_created_time(created_time: str) -> str:
    parsed_time = parse_created_time(created_time)
    if parsed_time:
        return parsed_time.strftime(EXCEL_CREATED_TIME_FORMAT)

    return str(created_time or "")


def build_post_link(post: dict) -> str:
    return f"https://www.facebook.com/{post.get('id')}"


def calculate_link_column_width(links: list[str]) -> int:
    longest_link = max((len(str(link or "")) for link in links), default=0)
    return min(EXCEL_MAX_COLUMN_WIDTH, max(LINK_COLUMN_MIN_WIDTH, longest_link + 4))


def build_group_posts_excel(
    group_id: str,
    posts: list[dict],
    include_group_column: bool = False,
) -> Path:
    export_dir = EXPORTS_DIR / "telegram"
    export_dir.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Posts"
    worksheet.freeze_panes = "A2"

    headers = ["Link bài", "Time created", "Message"]
    if include_group_column:
        headers = ["Nhóm"] + headers
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    thin_side = Side(style="thin", color="C9D2DB")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    body_font = Font(name=DEFAULT_FONT_NAME, size=DEFAULT_FONT_SIZE)
    header_font = Font(name=DEFAULT_FONT_NAME, size=DEFAULT_FONT_SIZE, bold=True)
    wrap_alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)
    no_wrap_alignment = Alignment(vertical="top", horizontal="left", wrap_text=False)
    link_column_index = 2 if include_group_column else 1
    link_column_letter = get_column_letter(link_column_index)
    post_links = [build_post_link(post) for post in posts]

    worksheet.append(headers)
    for column_index, cell in enumerate(worksheet[1], start=1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = no_wrap_alignment if column_index == link_column_index else wrap_alignment
        cell.border = border

    for post, post_link in zip(posts, post_links):
        row = [
            post_link,
            format_excel_created_time(post.get("created_time")),
            sanitize_excel_message(post.get("message")),
        ]
        if include_group_column:
            row = [str(post.get("group_id") or "")] + row
        worksheet.append(row)

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for column_index, cell in enumerate(row, start=1):
            cell.font = body_font
            cell.alignment = no_wrap_alignment if column_index == link_column_index else wrap_alignment
            cell.border = border

    if include_group_column:
        worksheet.column_dimensions["A"].width = 45
        worksheet.column_dimensions["B"].width = calculate_link_column_width(post_links)
        worksheet.column_dimensions["C"].width = 22
        worksheet.column_dimensions["D"].width = 120
    else:
        worksheet.column_dimensions["A"].width = calculate_link_column_width(post_links)
        worksheet.column_dimensions["B"].width = 22
        worksheet.column_dimensions["C"].width = 120

    worksheet.column_dimensions[link_column_letter].bestFit = True

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")

    # Main row parser.
    group_name = extract_group_name(group_id)
    group_name = safe_filename(group_name)

    file_path = export_dir / f"group_{group_name}_{timestamp}.xlsx"

    workbook.save(file_path)
    return file_path
